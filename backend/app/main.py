"""
GST Invoice Matching Agent — FastAPI backend.

Endpoints
---------
GET  /api/health                 -> liveness check
GET  /api/demo-data              -> regenerate + return the bundled synthetic dataset
POST /api/reconcile              -> upload gstr2b.csv + purchase_ledger.csv, get the full report
POST /api/reconcile/demo         -> run reconciliation on the bundled synthetic dataset (no upload needed)
GET  /                           -> serves the dashboard (frontend/index.html)

Run with:  uvicorn app.main:app --reload --app-dir backend
"""
import csv
import hashlib
import io
import logging
import os
import re
from typing import List, Dict

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from . import data_gen
from .matcher import reconcile
from .persistence import get_run, init_db, list_runs, save_run, update_exception
from .review import review_report

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # backend/
DATA_DIR = os.path.join(BASE_DIR, "data")
FRONTEND_DIR = os.path.join(os.path.dirname(BASE_DIR), "frontend")

app = FastAPI(title="GST Invoice Matching Agent", version="1.0.0")
logger = logging.getLogger(__name__)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
init_db()


REQUIRED_COLS = {"gstin", "supplier_name", "invoice_number", "invoice_date",
                 "taxable_value", "tax_amount", "total_value"}


def _parse_amount(value: str) -> float:
    cleaned = re.sub(r"[^0-9.\-]", "", str(value or ""))
    return float(cleaned)


def _read_csv_rows(raw_bytes: bytes, label: str) -> tuple[List[Dict], List[str]]:
    text = raw_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = {str(header).strip().lower(): header for header in (reader.fieldnames or [])}
    missing = REQUIRED_COLS - set(headers)
    if missing:
        raise HTTPException(400, f"{label} is missing required columns: {sorted(missing)}")
    rows = []
    warnings = []
    if any(str(header) != str(header).strip() or str(header) != str(header).lower() for header in (reader.fieldnames or [])):
        warnings.append(f"{label}: normalized whitespace/casing in column headers")
    for row_number, raw_row in enumerate(reader, start=2):
        row = {name: str(raw_row.get(source, "") or "").strip() for name, source in headers.items()}
        if None in raw_row:
            warnings.append(f"{label} row {row_number}: extra CSV values ignored; quote comma-containing fields")
            logger.warning("%s row %d had extra CSV values", label, row_number)
        if any(not row.get(column) for column in REQUIRED_COLS):
            warnings.append(f"{label} row {row_number} skipped: missing required field")
            logger.warning("%s row %d skipped: missing required field", label, row_number)
            continue
        try:
            for column in ("taxable_value", "tax_amount", "total_value"):
                original = row[column]
                row[column] = _parse_amount(row[column])
                if str(row[column]) != original:
                    warnings.append(f"{label} row {row_number}: cleaned {column} amount")
        except ValueError:
            warnings.append(f"{label} row {row_number} skipped: invalid amount")
            logger.warning("%s row %d skipped: invalid amount", label, row_number)
            continue
        rows.append(row)
    return rows, warnings


def _load_bundled_csv(filename: str) -> List[Dict]:
    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        os.makedirs(DATA_DIR, exist_ok=True)
        data_gen.generate(DATA_DIR)
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _file_hash(raw_bytes: bytes) -> str:
    return hashlib.sha256(raw_bytes).hexdigest()


def _persist(report: dict, gstr2b_bytes: bytes, ledger_bytes: bytes) -> dict:
    report["run_id"] = save_run(report, _file_hash(gstr2b_bytes), _file_hash(ledger_bytes))
    return report


@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.get("/api/demo-data")
def demo_data():
    os.makedirs(DATA_DIR, exist_ok=True)
    n2b, nledger = data_gen.generate(DATA_DIR)
    return {"gstr2b_rows": n2b, "ledger_rows": nledger, "message": "Synthetic data regenerated."}


@app.post("/api/reconcile/demo")
async def reconcile_demo(
    amount_tolerance_pct: float = Query(0.5, ge=0, le=10),
    fuzzy_threshold: float = Query(82, ge=50, le=100),
):
    gstr2b_path = os.path.join(DATA_DIR, "gstr2b.csv")
    ledger_path = os.path.join(DATA_DIR, "purchase_ledger.csv")
    gstr2b_rows = _load_bundled_csv("gstr2b.csv")
    ledger_rows = _load_bundled_csv("purchase_ledger.csv")
    if not gstr2b_rows or not ledger_rows:
        raise HTTPException(500, "Bundled demo data is missing or empty.")
    with open(gstr2b_path, "rb") as gstr2b_file, open(ledger_path, "rb") as ledger_file:
        report = await review_report(reconcile(
            gstr2b_rows,
            ledger_rows,
            amount_tolerance_pct=amount_tolerance_pct / 100,
            fuzzy_invoice_threshold=fuzzy_threshold / 100,
        ))
        return _persist(report, gstr2b_file.read(), ledger_file.read())


@app.post("/api/reconcile")
async def reconcile_upload(
    gstr2b_file: UploadFile = File(...),
    ledger_file: UploadFile = File(...),
    amount_tolerance_pct: float = Form(0.5),
    fuzzy_threshold: float = Form(82),
):
    gstr2b_bytes = await gstr2b_file.read()
    ledger_bytes = await ledger_file.read()
    try:
        gstr2b_rows, gstr2b_warnings = _read_csv_rows(gstr2b_bytes, "gstr2b_file")
        ledger_rows, ledger_warnings = _read_csv_rows(ledger_bytes, "ledger_file")
    except Exception as e:
        if isinstance(e, HTTPException):
            raise
        raise HTTPException(400, f"Could not parse uploaded CSVs: {e}")
    if not 0 <= amount_tolerance_pct <= 10 or not 50 <= fuzzy_threshold <= 100:
        raise HTTPException(400, "Matching settings are outside the supported range.")
    report = await review_report(reconcile(
        gstr2b_rows,
        ledger_rows,
        amount_tolerance_pct=amount_tolerance_pct / 100,
        fuzzy_invoice_threshold=fuzzy_threshold / 100,
    ))
    report["data_quality_warnings"] = gstr2b_warnings + ledger_warnings
    return _persist(report, gstr2b_bytes, ledger_bytes)


@app.get("/api/runs")
def runs():
    return {"runs": list_runs()}


@app.get("/api/runs/{run_id}")
def run_detail(run_id: int):
    result = get_run(run_id)
    if result is None:
        raise HTTPException(404, "Reconciliation run not found.")
    return result


@app.patch("/api/runs/{run_id}/exceptions/{exception_id}")
def resolve_exception(run_id: int, exception_id: int, payload: Dict):
    try:
        result = update_exception(
            run_id, exception_id, str(payload.get("status", "")).upper(), payload.get("note")
        )
    except ValueError as error:
        raise HTTPException(400, str(error))
    if result is None:
        raise HTTPException(404, "Exception not found for this reconciliation run.")
    return result


@app.post("/api/reconcile/demo/narrate")
async def narrate(report: Dict):
    import httpx

    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise HTTPException(503, "Set GEMINI_API_KEY to enable narration.")
    patterns = sorted(report.get("counts_by_status", {}).items(), key=lambda item: item[1], reverse=True)[:3]
    prompt = ("Write 2-3 plain-English paragraphs for a non-technical business owner. "
              f"Mention total ITC at risk of ₹{report.get('itc_at_risk_amount', 0):,.2f}, "
              f"the top 3 exception patterns {patterns}, and one recommended action. "
              "Do not invent figures or use technical implementation language.\n"
              f"Report: {report}")
    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent"
    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.post(f"{url}?key={api_key}", json={"contents": [{"parts": [{"text": prompt}]}]})
    if response.is_error:
        raise HTTPException(502, "Gemini narration request failed.")
    body = response.json()
    return {"summary": body["candidates"][0]["content"]["parts"][0]["text"]}


@app.get("/api/reconcile/demo/export")
def export_demo():
    gstr2b_rows = _load_bundled_csv("gstr2b.csv")
    ledger_rows = _load_bundled_csv("purchase_ledger.csv")
    report = reconcile(gstr2b_rows, ledger_rows)
    tax_by_invoice = {(row["gstin"], row["invoice_number"]): float(row["tax_amount"]) for row in gstr2b_rows}
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exceptions"
    headers = ["Status", "GSTIN", "Supplier", "2B Invoice", "Ledger Invoice", "2B Taxable",
               "Ledger Taxable", "ITC At Risk", "Confidence", "Reason"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    rows = sorted((row for row in report["results"] if row["status"] != "MATCHED"), key=lambda row: row["status"])
    for row in rows:
        invoice = row["gstr2b_invoice"] or row["ledger_invoice"]
        risk = tax_by_invoice.get((row["gstin"], invoice), 0) if row["status"] == "MISSING_IN_LEDGER" else 0
        sheet.append([row["status"], row["gstin"], row["supplier_name"], row["gstr2b_invoice"],
                      row["ledger_invoice"], row["taxable_value_2b"], row["taxable_value_ledger"],
                      risk, row["confidence"], row["reason"]])
        if risk:
            sheet.cell(sheet.max_row, 8).fill = PatternFill("solid", fgColor="FFF2CC")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=exception-report.xlsx"})


# Serve the dashboard as static files (index.html at "/")
if os.path.isdir(FRONTEND_DIR):
    app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="frontend")
